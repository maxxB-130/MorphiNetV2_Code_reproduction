"""
XLSX export functionality for MorphiNet comprehensive metrics reporting.

This module handles the export of testing metrics to Excel format following 
the TestBench structure with multi-level indexing for methods and metrics.
"""

import os
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Any
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

__all__ = [
    "MetricsExporter", 
    "export_metrics_to_xlsx",
    "export_ablation_study_to_xlsx",
    "export_mr_ablation_to_xlsx", 
    "export_ct_ablation_to_xlsx",
    "create_ablation_comparison_dataframe"
]

class MetricsExporter:
    """
    Handles comprehensive metric export to XLSX format compatible with TestBench structure.
    
    Creates multi-level indexed DataFrames with:
    - Index Level 0: Method names (e.g., 'MorphiNet') 
    - Index Level 1: Metric names (e.g., 'Dice', 'Hausdorff', etc.)
    - Columns: Case IDs
    - Sheets: Anatomical parts (e.g., 'lv', 'rv', 'myo')
    """
    
    METRIC_ORDER = [
        'Dice', 'Hausdorff', 'Chamfer (mm)', 'ASD (mm)', 
        'Aspect Ratio', 'Skew', 'Jacobian', 'Jacobian Ratio < 0.7',
        'Mean Normal Consistency', 'Non-manifold Face Ratio'
    ]
    
    def __init__(self, method_name: str = "MorphiNet"):
        """
        Initialize exporter.
        
        Args:
            method_name: Name of the method being evaluated (e.g., 'MorphiNet')
        """
        self.method_name = method_name
        self.metrics_data = defaultdict(lambda: defaultdict(list))
        self.case_ids = []
        
    def add_batch_metrics(self, case_ids: List[str], metrics: Dict[str, Any]):
        """
        Add metrics from a batch of test cases.
        
        Args:
            case_ids: List of case identifiers for this batch
            metrics: Dictionary containing all computed metrics
                    Structure: {
                        'dice_scores': {'LV': [values], 'RV': [values], 'MYO': [values]},
                        'hausdorff_scores': {...},
                        'mesh_metrics': {
                            'LV': {'asd': [values], 'aspect_ratio': [values], ...},
                            ...
                        }
                    }
        """
        self.case_ids.extend(case_ids)
        
        # Extract and organize all metrics by anatomical part
        for case_idx, case_id in enumerate(case_ids):
            # Standard MONAI metrics (dice, hausdorff)
            if 'dice_scores' in metrics:
                for part, values in metrics['dice_scores'].items():
                    if case_idx < len(values):
                        self.metrics_data[part.upper()]['Dice'].append(values[case_idx])
                    
            if 'hausdorff_scores' in metrics:
                for part, values in metrics['hausdorff_scores'].items():
                    if case_idx < len(values):
                        self.metrics_data[part.upper()]['Hausdorff'].append(values[case_idx])
            
            # Advanced mesh metrics
            if 'mesh_metrics' in metrics:
                mesh_metrics = metrics['mesh_metrics']
                for part in mesh_metrics:
                    part_upper = part.upper()
                    part_data = mesh_metrics[part]
                    
                    # Map metric names to standard XLSX column names
                    metric_mapping = {
                        'chamfer': 'Chamfer (mm)',
                        'asd': 'ASD (mm)',
                        'aspect_ratio': 'Aspect Ratio',
                        'skew': 'Skew',
                        'jacobian': 'Jacobian',
                        'jacobian_ratio_low': 'Jacobian Ratio < 0.7',
                        'normal_consistency': 'Mean Normal Consistency',
                        'nm_face_ratio': 'Non-manifold Face Ratio'
                    }
                    
                    for metric_key, xlsx_name in metric_mapping.items():
                        if metric_key in part_data and case_idx < len(part_data[metric_key]):
                            self.metrics_data[part_upper][xlsx_name].append(part_data[metric_key][case_idx])
    
    def create_dataframes(self) -> Dict[str, pd.DataFrame]:
        """
        Create pandas DataFrames with multi-level indexing for each anatomical part.
        
        Returns:
            Dictionary mapping part names to DataFrames
        """
        dataframes = {}
        
        for part, part_metrics in self.metrics_data.items():
            # Create multi-level index structure
            methods = []
            metrics = []
            values = []
            
            for metric_name in self.METRIC_ORDER:
                if metric_name in part_metrics and len(part_metrics[metric_name]) > 0:
                    methods.append(self.method_name)
                    metrics.append(metric_name)
                    # Pad values to match case_ids length
                    metric_values = part_metrics[metric_name]
                    while len(metric_values) < len(self.case_ids):
                        metric_values.append(np.nan)
                    values.append(metric_values[:len(self.case_ids)])
            
            if values:  # Only create DataFrame if we have data
                # Create DataFrame with multi-level index
                df = pd.DataFrame(
                    values,
                    index=pd.MultiIndex.from_arrays([methods, metrics], names=['Method', 'Metric']),
                    columns=self.case_ids,
                    dtype=object
                )
                
                # Transpose to match TestBench format (case_ids as rows)
                df = df.transpose()
                dataframes[part.lower()] = df
                
        return dataframes
    
    def export_to_xlsx(self, filepath: str):
        """
        Export all metrics to XLSX file.
        
        Args:
            filepath: Output file path (should end with .xlsx)
        """
        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        
        dataframes = self.create_dataframes()
        
        if not dataframes:
            print(f"Warning: No metrics data to export to {filepath}")
            return
            
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            for part_name, df in dataframes.items():
                df.to_excel(writer, sheet_name=part_name)
                print(f"Exported {part_name} metrics: {df.shape}")
        

def export_metrics_to_xlsx(
    case_ids: List[str], 
    dice_scores: Dict[str, List[float]], 
    hausdorff_scores: Dict[str, List[float]], 
    mesh_metrics: Dict[str, Dict[str, np.ndarray]], 
    output_path: str,
    method_name: str = "MorphiNet"
) -> None:
    """
    Convenience function to export metrics directly to XLSX.
    
    Args:
        case_ids: List of test case identifiers
        dice_scores: Dictionary of Dice scores by anatomical part  
        hausdorff_scores: Dictionary of Hausdorff distances by anatomical part
        mesh_metrics: Dictionary of advanced mesh metrics by anatomical part
        output_path: Output XLSX file path
        method_name: Method name for indexing
    """
    exporter = MetricsExporter(method_name)
    
    # Organize metrics into expected structure
    organized_metrics = {
        'dice_scores': dice_scores,
        'hausdorff_scores': hausdorff_scores,
        'mesh_metrics': mesh_metrics
    }
    
    exporter.add_batch_metrics(case_ids, organized_metrics)
    exporter.export_to_xlsx(output_path)


def validate_metrics_structure(metrics: Dict[str, Any]) -> bool:
    """
    Validate that metrics dictionary has expected structure.
    
    Args:
        metrics: Metrics dictionary to validate
        
    Returns:
        True if structure is valid, False otherwise
    """
    required_keys = ['dice_scores', 'hausdorff_scores', 'mesh_metrics']
    
    for key in required_keys:
        if key not in metrics:
            print(f"Warning: Missing required key '{key}' in metrics")
            return False
    
    # Validate mesh_metrics structure
    if 'mesh_metrics' in metrics:
        mesh_metrics = metrics['mesh_metrics']
        required_mesh_keys = ['asd', 'aspect_ratio', 'skew', 'jacobian', 
                             'normal_consistency', 'nm_face_ratio', 'chamfer']
        
        for part in mesh_metrics:
            for key in required_mesh_keys:
                if key not in mesh_metrics[part]:
                    print(f"Warning: Missing mesh metric '{key}' for part '{part}'")
                    return False
    
    return True


def export_ablation_study_to_xlsx(ablation_metrics, export_path, dataset_export_name):
    """
    Export ablation study metrics to XLSX files, creating separate files for MR and CT.
    
    Args:
        ablation_metrics: Dictionary containing ablation metrics data
        export_path: Output file path for the XLSX export
        dataset_export_name: Dataset name for file naming
    """
    modal = ablation_metrics['modal']
    dataset_name = ablation_metrics['dataset_name']
    
    if modal == 'mr':
        export_mr_ablation_to_xlsx(ablation_metrics, export_path, dataset_export_name)
    elif modal == 'ct':
        export_ct_ablation_to_xlsx(ablation_metrics, export_path, dataset_export_name)
    else:
        print(f"Warning: Unknown modality '{modal}' for ablation study export")


def export_mr_ablation_to_xlsx(ablation_metrics, export_path, dataset_export_name):
    """
    Export MR ablation study metrics with ED/ES phase separation and proper nested column structure.
    
    Args:
        ablation_metrics: Dictionary containing MR ablation metrics with enhanced structure
        export_path: Output file path for the XLSX export
        dataset_export_name: Dataset name for file naming
    """
    case_ids = ablation_metrics['case_ids']
    phases = ablation_metrics['phases']
    before_dice = ablation_metrics['before_resnet_dice']
    after_dice = ablation_metrics['after_resnet_dice']
    before_hausdorff = ablation_metrics['before_resnet_hausdorff']
    after_hausdorff = ablation_metrics['after_resnet_hausdorff']
    volume_diffs = ablation_metrics['volume_differences']
    
    # Create workbook and worksheets
    wb = Workbook()
    ws = wb.active
    ws.title = "Case_Details"
    
    # Define column structure for proper nesting
    data_columns = ['Case_ID', 'Phase']
    header_structure = []
    
    # Add metadata columns
    header_structure.extend([
        ('Metadata', 'Case_ID', ''),
        ('Metadata', 'Phase', '')
    ])
    
    # Add metric columns
    for metric_type in ['Dice', 'Hausdorff', 'Volume_Diff']:
        for label in ['LV', 'MYO', 'RV']:
            if metric_type == 'Volume_Diff':
                measures = ['Absolute', 'Percentage', 'Before_Count', 'After_Count']
            else:
                measures = ['Before', 'After']
            
            for measure in measures:
                data_columns.append(f'{metric_type}_{label}_{measure}')
                header_structure.append((metric_type, label, measure))
    
    # Prepare data rows
    data_rows = []
    for i, case_id in enumerate(case_ids):
        phase = phases[i] if i < len(phases) else 'unknown'
        row = [case_id, phase]
        
        # Add data for each metric type and label
        for metric_type in ['Dice', 'Hausdorff', 'Volume_Diff']:
            for label in ['LV', 'MYO', 'RV']:
                if metric_type == 'Dice':
                    before_val = before_dice[label][i] if i < len(before_dice[label]) else 0.0
                    after_val = after_dice[label][i] if i < len(after_dice[label]) else 0.0
                    row.extend([before_val, after_val])
                elif metric_type == 'Hausdorff':
                    before_val = before_hausdorff[label][i] if i < len(before_hausdorff[label]) else 0.0
                    after_val = after_hausdorff[label][i] if i < len(after_hausdorff[label]) else 0.0
                    row.extend([before_val, after_val])
                elif metric_type == 'Volume_Diff':
                    vol_diff = volume_diffs[label][i] if i < len(volume_diffs[label]) else {}
                    row.extend([
                        vol_diff.get('absolute_diff', 0),
                        vol_diff.get('percentage_diff', 0.0),
                        vol_diff.get('before_count', 0),
                        vol_diff.get('after_count', 0)
                    ])
        
        data_rows.append(row)
    
    # Create nested headers using openpyxl
    _create_nested_headers(ws, header_structure)
    
    # Add data starting from row 4 (after headers)
    for row_idx, row_data in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Phase_Summary")
    _create_mr_summary_sheet(ws_summary, case_ids, phases, before_dice, after_dice, 
                           before_hausdorff, after_hausdorff)
    
    # Save workbook
    output_path = os.path.join(export_path, f'ablation_MR_{dataset_export_name}.xlsx')
    wb.save(output_path)
    
    print(f"MR ablation study exported to: {output_path}")
    print(f"Total cases: {len(data_rows)}, ED: {len([p for p in phases if p == 'ED'])}, ES: {len([p for p in phases if p == 'ES'])}")


def export_ct_ablation_to_xlsx(ablation_metrics, export_path, dataset_export_name):
    """
    Export CT ablation study metrics with proper nested column structure.
    
    Args:
        ablation_metrics: Dictionary containing CT ablation metrics with enhanced structure
        export_path: Output file path for the XLSX export
        dataset_export_name: Dataset name for file naming
    """
    case_ids = ablation_metrics['case_ids']
    before_dice = ablation_metrics['before_resnet_dice']
    after_dice = ablation_metrics['after_resnet_dice']
    before_hausdorff = ablation_metrics['before_resnet_hausdorff']
    after_hausdorff = ablation_metrics['after_resnet_hausdorff']
    volume_diffs = ablation_metrics['volume_differences']
    
    # Create workbook and worksheets
    wb = Workbook()
    ws = wb.active
    ws.title = "Case_Details"
    
    # Define column structure for proper nesting
    data_columns = ['Case_ID']
    header_structure = []
    
    # Add metadata columns
    header_structure.append(('Metadata', 'Case_ID', ''))
    
    # Add metric columns
    for metric_type in ['Dice', 'Hausdorff', 'Volume_Diff']:
        for label in ['LV', 'MYO', 'RV']:
            if metric_type == 'Volume_Diff':
                measures = ['Absolute', 'Percentage', 'Before_Count', 'After_Count']
            else:
                measures = ['Before', 'After']
            
            for measure in measures:
                data_columns.append(f'{metric_type}_{label}_{measure}')
                header_structure.append((metric_type, label, measure))
    
    # Prepare data rows
    data_rows = []
    for i, case_id in enumerate(case_ids):
        row = [case_id]
        
        # Add data for each metric type and label
        for metric_type in ['Dice', 'Hausdorff', 'Volume_Diff']:
            for label in ['LV', 'MYO', 'RV']:
                if metric_type == 'Dice':
                    before_val = before_dice[label][i] if i < len(before_dice[label]) else 0.0
                    after_val = after_dice[label][i] if i < len(after_dice[label]) else 0.0
                    row.extend([before_val, after_val])
                elif metric_type == 'Hausdorff':
                    before_val = before_hausdorff[label][i] if i < len(before_hausdorff[label]) else 0.0
                    after_val = after_hausdorff[label][i] if i < len(after_hausdorff[label]) else 0.0
                    row.extend([before_val, after_val])
                elif metric_type == 'Volume_Diff':
                    vol_diff = volume_diffs[label][i] if i < len(volume_diffs[label]) else {}
                    row.extend([
                        vol_diff.get('absolute_diff', 0),
                        vol_diff.get('percentage_diff', 0.0),
                        vol_diff.get('before_count', 0),
                        vol_diff.get('after_count', 0)
                    ])
        
        data_rows.append(row)
    
    # Create nested headers using openpyxl
    _create_nested_headers(ws, header_structure)
    
    # Add data starting from row 4 (after headers)
    for row_idx, row_data in enumerate(data_rows, start=4):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Create summary sheet
    ws_summary = wb.create_sheet("Summary")
    _create_ct_summary_sheet(ws_summary, case_ids, before_dice, after_dice, 
                           before_hausdorff, after_hausdorff)
    
    # Save workbook
    output_path = os.path.join(export_path, f'ablation_CT_{dataset_export_name}.xlsx')
    wb.save(output_path)
    
    print(f"CT ablation study exported to: {output_path}")
    print(f"Total cases: {len(data_rows)}")


def create_ablation_comparison_dataframe(case_ids, before_scores, after_scores, phases=None):
    """
    Create a structured DataFrame for ablation study comparison.
    
    Args:
        case_ids: List of case identifiers
        before_scores: List of Dice scores before ResNet
        after_scores: List of Dice scores after ResNet
        phases: Optional list of cardiac phases (for MR data)
        
    Returns:
        pd.DataFrame: Structured comparison data
    """
    data = {
        'Case_ID': case_ids,
        'Before_ResNet': before_scores,
        'After_ResNet': after_scores,
        'Improvement': [after - before for before, after in zip(before_scores, after_scores)]
    }
    
    if phases is not None:
        data['Phase'] = phases
    
    return pd.DataFrame(data)


def _create_nested_headers(ws, header_structure):
    """
    Create nested headers with merged cells in an Excel worksheet.
    
    Args:
        ws: openpyxl worksheet
        header_structure: List of tuples (metric_type, anatomy, measure)
    """
    # Define styles
    header_style = Font(bold=True, size=10)
    metadata_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    dice_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    hausdorff_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    volume_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Row 1: Metric Type headers
    current_col = 1
    metric_groups = {}
    
    for metric_type, anatomy, measure in header_structure:
        if metric_type not in metric_groups:
            metric_groups[metric_type] = {'start': current_col, 'count': 0}
        metric_groups[metric_type]['count'] += 1
        current_col += 1
    
    # Merge cells for metric types in row 1
    col_idx = 1
    for metric_type, info in metric_groups.items():
        start_col = info['start']
        end_col = start_col + info['count'] - 1
        
        if start_col == end_col:
            cell = ws.cell(row=1, column=start_col, value=metric_type)
        else:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=metric_type)
        
        cell.font = header_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Apply colors based on metric type
        if metric_type == 'Metadata':
            cell.fill = metadata_fill
        elif metric_type == 'Dice':
            cell.fill = dice_fill
        elif metric_type == 'Hausdorff':
            cell.fill = hausdorff_fill
        elif metric_type == 'Volume_Diff':
            cell.fill = volume_fill
    
    # Row 2: Anatomy headers
    anatomy_groups = {}
    current_col = 1
    
    for metric_type, anatomy, measure in header_structure:
        key = f"{metric_type}_{anatomy}"
        if key not in anatomy_groups:
            anatomy_groups[key] = {'start': current_col, 'count': 0, 'anatomy': anatomy, 'metric_type': metric_type}
        anatomy_groups[key]['count'] += 1
        current_col += 1
    
    # Merge cells for anatomy in row 2
    for key, info in anatomy_groups.items():
        start_col = info['start']
        end_col = start_col + info['count'] - 1
        anatomy = info['anatomy']
        metric_type = info['metric_type']
        
        if start_col == end_col:
            cell = ws.cell(row=2, column=start_col, value=anatomy)
        else:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
            cell = ws.cell(row=2, column=start_col, value=anatomy)
        
        cell.font = header_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Apply same colors as parent metric type
        if metric_type == 'Metadata':
            cell.fill = metadata_fill
        elif metric_type == 'Dice':
            cell.fill = dice_fill
        elif metric_type == 'Hausdorff':
            cell.fill = hausdorff_fill
        elif metric_type == 'Volume_Diff':
            cell.fill = volume_fill
    
    # Row 3: Measure headers
    for col_idx, (metric_type, anatomy, measure) in enumerate(header_structure, start=1):
        cell = ws.cell(row=3, column=col_idx, value=measure if measure else anatomy)
        cell.font = header_style
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
        
        # Apply same colors as parent metric type
        if metric_type == 'Metadata':
            cell.fill = metadata_fill
        elif metric_type == 'Dice':
            cell.fill = dice_fill
        elif metric_type == 'Hausdorff':
            cell.fill = hausdorff_fill
        elif metric_type == 'Volume_Diff':
            cell.fill = volume_fill
    
    # Auto-adjust column widths
    for col_idx in range(1, len(header_structure) + 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 12


def _create_mr_summary_sheet(ws, case_ids, phases, before_dice, after_dice, before_hausdorff, after_hausdorff):
    """Create summary sheet for MR ablation study."""
    ws.cell(row=1, column=1, value="Phase Summary Statistics").font = Font(bold=True, size=14)
    
    # Create headers
    headers = ['Phase', 'N_Cases']
    for metric_type in ['Dice', 'Hausdorff']:
        for label in ['LV', 'MYO', 'RV']:
            headers.extend([
                f'{metric_type}_{label}_Before_Mean',
                f'{metric_type}_{label}_Before_Std',
                f'{metric_type}_{label}_After_Mean',
                f'{metric_type}_{label}_After_Std'
            ])
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Calculate statistics by phase
    row_idx = 4
    for phase in ['ED', 'ES', 'unknown']:
        phase_indices = [i for i, p in enumerate(phases) if p == phase]
        if not phase_indices:
            continue
            
        row_data = [phase, len(phase_indices)]
        
        for metric_type in ['Dice', 'Hausdorff']:
            for label in ['LV', 'MYO', 'RV']:
                if metric_type == 'Dice':
                    before_values = [before_dice[label][i] for i in phase_indices if i < len(before_dice[label])]
                    after_values = [after_dice[label][i] for i in phase_indices if i < len(after_dice[label])]
                else:
                    before_values = [before_hausdorff[label][i] for i in phase_indices if i < len(before_hausdorff[label])]
                    after_values = [after_hausdorff[label][i] for i in phase_indices if i < len(after_hausdorff[label])]
                
                before_mean = np.mean(before_values) if before_values else 0.0
                before_std = np.std(before_values) if before_values else 0.0
                after_mean = np.mean(after_values) if after_values else 0.0
                after_std = np.std(after_values) if after_values else 0.0
                
                row_data.extend([before_mean, before_std, after_mean, after_std])
        
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1


def _create_ct_summary_sheet(ws, case_ids, before_dice, after_dice, before_hausdorff, after_hausdorff):
    """Create summary sheet for CT ablation study."""
    ws.cell(row=1, column=1, value="Overall Summary Statistics").font = Font(bold=True, size=14)
    
    # Create headers
    headers = ['Group', 'N_Cases']
    for metric_type in ['Dice', 'Hausdorff']:
        for label in ['LV', 'MYO', 'RV']:
            headers.extend([
                f'{metric_type}_{label}_Before_Mean',
                f'{metric_type}_{label}_Before_Std',
                f'{metric_type}_{label}_After_Mean',
                f'{metric_type}_{label}_After_Std'
            ])
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Calculate overall statistics
    row_data = ['Overall', len(case_ids)]
    
    for metric_type in ['Dice', 'Hausdorff']:
        for label in ['LV', 'MYO', 'RV']:
            if metric_type == 'Dice':
                before_values = before_dice[label]
                after_values = after_dice[label]
            else:
                before_values = before_hausdorff[label]
                after_values = after_hausdorff[label]
            
            before_mean = np.mean(before_values) if before_values else 0.0
            before_std = np.std(before_values) if before_values else 0.0
            after_mean = np.mean(after_values) if after_values else 0.0
            after_std = np.std(after_values) if after_values else 0.0
            
            row_data.extend([before_mean, before_std, after_mean, after_std])
    
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=4, column=col_idx, value=value)